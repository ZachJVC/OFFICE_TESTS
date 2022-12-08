#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
Created on Mon Nov 16 14:15:59 2020

ballReportGenerator.py
    ____        ____   ____                        __ 
   / __ )____ _/ / /  / __ \___  ____  ____  _____/ /_
  / __  / __ `/ / /  / /_/ / _ \/ __ \/ __ \/ ___/ __/
 / /_/ / /_/ / / /  / _, _/  __/ /_/ / /_/ / /  / /_  
/_____/\__,_/_/_/  /_/ |_|\___/ .___/\____/_/   \__/  
  / ____/__  ____  ___  _____/_/_ _/ /_____  _____    
 / / __/ _ \/ __ \/ _ \/ ___/ __ `/ __/ __ \/ ___/    
/ /_/ /  __/ / / /  __/ /  / /_/ / /_/ /_/ / /        
\____/\___/_/ /_/\___/_/   \__,_/\__/\____/_/         
                                                      

@author: sportable
"""

#%% Import libraries

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from tqdm import tqdm
import sys
import os
import datetime
import src.productionConfig as configReader
import subprocess
from src.datatypes import testStage

#%% Define class
class ballReportGenerator():
      
    def __init__(self, stage, dataOrder = "anti-chronological"):
        
        # USER SETTINGS
        if stage == testStage.VALVE :
            self.mainSheetName = f'{testStage.VALVE} Report'
            self.testLogFolder = f"@{testStage.VALVE}"
        elif stage == testStage.BLADDER :
            self.mainSheetName = f'{testStage.BLADDER} Report'
            self.testLogFolder = f"@{testStage.BLADDER}"
        elif stage == testStage.BALL :
            self.mainSheetName = f'{testStage.BALL} Report'
            self.testLogFolder = f"@{testStage.BALL}"
        
        # Set the data publishing order in the report
        self.dataOrder = dataOrder
        
        # Set Excel conditional formatting colours
        self.redFill = PatternFill(start_color='FF7171', end_color='FF7171', fill_type='solid')
        self.yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.greenFill = PatternFill(start_color='75FF75', end_color='75FF75', fill_type='solid')

        # Get the path, parent directory and folder names in the folder related to the test currently run
        self.path = os.getcwd()
        if self.path.split('/')[-1] == 'release' :
            self.parent = self.path.split('/')[-2] 
        else :
            self.parent = self.path.split('/')[-1]
        self.folders = os.listdir(self.testLogFolder)

        # If there are any issues with a testlog, we will log issues here
        self.logsToInvestigate = datetime.datetime.now().strftime("%Y%m%d") + '_Problematic_Logs' + '.txt'
        
        # Set name for output excel file
        self.reportSummary  = self.parent + '_Electrical_Test_Report.xlsx'
        
        # Check if the results file already exists
        self.reportSummaryExists =  os.path.isfile(self.reportSummary)              

        # Read in swiftPassFailCriteria.xlsx
        self.setPassFailCriteria()        
                
        # Initialise these containers to be initially empty         
        self.testedDevices  = []
        self.parsedLogs     = []
        
        # If it exists get the devices already tested and find first empty row
        if self.reportSummaryExists:

            workbook = load_workbook(self.reportSummary)
            if self.mainSheetName not in workbook.sheetnames :
                self.__initialiseReportSummary(self.reportSummaryExists)
                self.rowToWriteData = 2
            else:
                self.getTestedDevicesFromReport()
                self.getParsedLogFilesFromReport()
                self.getFirstEmptyRowFromReport()

            workbook.close()

        # If not initialise the Excel spreadsheet
        elif not self.reportSummaryExists:            
            
            self.__initialiseReportSummary(self.reportSummaryExists)

            # Since it doesn't exist, we can write to the first row after the header
            self.rowToWriteData = 2                        

    def __extractTestDetails(self, dataframe):
        
        dataArray   = dataframe.values   # Convert the dataframe to an array
        keys        = dataArray[2::,0]
        values      = dataArray[2::,1]
        criteria    = dataArray[2::,2:4] # Takes columns 2 and 3
        
        self.testNames      = []
        self.testCriteria   = []    # Will contain a list of lists. If an entry is populated, contains "min" and "max"
        self.sectionHeaders = []
        self.numberOfTests  = len([i for i in criteria[:,0] if isinstance(i, str) or (isinstance(i, (int, float)) and not np.isnan(i))])
        
        # Enumerate over the strings in column 1 of passFailCriteria.xlsx
        for j, key in enumerate(keys):            
            if isinstance(key, str): # If it's a string then could be a test name
                
                # If it satisfies this then it must be a test name
                if isinstance(values[j], str) or (isinstance(values[j], (float, int)) and not np.isnan(values[j])):
                    
                    # Now we should check that this test is not an accidental duplicate (a user copying a row in Excel)
                    if key in self.testNames:
                        continue
                    else:
                        self.testNames.append(key)
           
                    if isinstance(criteria[j,0], (str, float, int)):
                        self.testCriteria.append(list(criteria[j,:]))
                    elif np.isnan(criteria[j,0]):
                        self.testCriteria.append(list([np.nan, np.nan]))
                 
                # Otherwise it's a section header
                else:
                    self.sectionHeaders.append(key)
            else:
                None
        
    def __initialiseReportSummary(self, reportExists = False):
        """ Writes the header to Report Summary"""
        
        columns = ['Master Pass/Fail','Retest?'] + self.testNames + ['Log_File_Name']        
        header = pd.DataFrame(columns = columns)
        if reportExists == True :
            with pd.ExcelWriter(self.reportSummary, engine="openpyxl", mode='a') as writer :
                header.to_excel(writer, sheet_name=self.mainSheetName, index=False) 
        else :
            header.to_excel(self.reportSummary, sheet_name=self.mainSheetName, index=False, engine='openpyxl')
        self.logFileNameColumn = len(columns)

        # Add the pass/fail criteria to a new tab
        self.writePassFailCriteriaToExcel()

    def createReportFor(self, file):
         
        deviceTestValues    = np.empty((1, len(self.testNames)+2), dtype=object)
        devicePassFail      = np.empty((1, len(self.testNames)+2), dtype=object)        
        testsInLogFile      = []
        deviceId            = []
        
        self.logData = open(file, 'r').read()
            
        # Go over the log line by line
        for line in self.logData.splitlines():
            
            # If the line contains a comma separation then may be a test/result pair
            if len(line.split(',')) == 2:
                testName        = line.split(',')[0]; testsInLogFile.append(testName)                
                testOutcome     = line.split(',')[1]                

                if testName == 'SERIAL_NUMBER':
                    deviceId = testOutcome                    
                    
                # It's possible that a test in the log doesn't exist in the passFailCriteria.xlsx
                # If so, catch the error and log it
                try:                    
                    testIndex       = self.testNames.index(testName)
                except:
                    # The test exists in the log file but not in the passFailCriteria.xlsx document                    
                    writeString = f"{file} --> Test {testName} DNE in ballPassFailCriteria.xlsx \n"
                    self.writeToLogsToInvestigate(writeString)                    
                    # Continue to next test
                    continue
                
                # The test does exist, get its position in self.testCriteria
                testCriteria    = self.testCriteria[testIndex]
                        
                # Store the device testing outcome
                deviceTestValues[0, testIndex+2] = testOutcome
                
                # determine if the device passes / fails this test
                devicePassFail[0, testIndex+2] = self.determinePassFail(testName, testIndex, testCriteria, testOutcome, file)
      
        
        # Once we've gone over every line in the logfile, we should check if any tests
        # in passFailCriteria.xlsx were not performed
        testsNotPerformed       = [test for test in self.testNames if test not in testsInLogFile]
        testIndsNotPerformed  = [self.testNames.index(test) for test in testsNotPerformed]
        
        if len(testsNotPerformed) != 0:
            for test in testsNotPerformed:
                writeString = f"{file} --> The test {test} which is in ballPassFailCriteria.xlsx was not performed\n"
                self.writeToLogsToInvestigate(writeString)

        # It can only pass if every test has been performed AND it has passed all of them
        if np.sum(devicePassFail == 'pass') == self.numberOfTests:
            devicePassFail[0, 0]   = 'pass'
            devicePassFail[0, 1]   = '-'
            
            deviceTestValues[0, 0] = 'pass'
        
        # Otherwise it fails
        else: 
            devicePassFail[0, 0]   = 'fail'
            devicePassFail[0, 1]   = '-'
            
            deviceTestValues[0, 0] = 'fail'

        # Check if it's a retest
        if pd.isnull(deviceId) or deviceId == "" :
            writeString = f"The device ID was not found in {file}\n"
            self.writeToLogsToInvestigate(writeString)            
            deviceTestValues[0, 1] = 'unknown'
        else:
            if deviceId in self.testedDevices:
                deviceTestValues[0, 1] = 'RETEST'
            else:
                deviceTestValues[0, 1] = ''            
        
        # Now that we've checked for a retest, we can add the device to self.testedDevices
        self.testedDevices.append(deviceId)

        columns = ['Master Pass/Fail','Retest?'] + self.testNames                 
        self.writeDataToExcel(deviceTestValues, devicePassFail, columns, file)
        self.highlightNonPerformedTestCellsRed(testIndsNotPerformed)

        return devicePassFail[0, 0]

    def determinePassFail(self, testName, testIndex, testCriteria, testOutcome, file):
        """ Returns the outcome of a test given a result and the criteria. Logs observed issues to self.logsToInvestigate"""
        
        # If the test criterion is a string, then the testOutcome should also be string. If not it's an error
        if isinstance(testCriteria[0], str):
            try:
                if testOutcome.lower() == 'pass' or testOutcome.lower() == 'ok' or testOutcome.lower() == 'complete':                    
                    testResult = 'pass'
                elif testOutcome == testCriteria[0] or testOutcome == testCriteria[1] :
                    testResult = 'pass'
                elif testOutcome.lower() == 'fail':
                    testResult = 'fail'
                else :
                    testResult = 'fail'
            except:
                testResult = 'ERROR'
                
                # The test outcome must be string if the test criterion is string
                with open(self.logsToInvestigate, "a") as resultFile:
                    writeString = f"{file} --> Test outcome of {testName} is not string, but test criterion in ballPassFailCriteria.xlsx is string\n"
                    print('\n\n' + writeString + '\n\n')
                    resultFile.write(writeString)
               
        elif isinstance(testCriteria[0], (int, float)):
            
            # If the criterion is nan then it's not a test
            if np.isnan(testCriteria[0]):
                testResult = '-'

             # If the test criterion is int/float (& not nan), then the testOutcome should also be int/float. If not it's an error
            else:
                try:
                    if testCriteria[0] <= float(testOutcome) <= testCriteria[1]:
                        testResult = 'pass'
                    else:
                        testResult = 'fail'
                except:
                    testResult = 'ERROR'
                    
                    # The test outcome must be string if the test criterion is string
                    with open(self.logsToInvestigate, "a") as resultFile:
                        writeString = f"Test outcome of {testName} in {file} is not int/flaot, but test criterion in ballPassFailCriteria.xlsx is int/float\n"
                        print('\n\n' + writeString + '\n\n')
                        resultFile.write(writeString)

        else:
            # There's no test, so indicate this with a "-"
            testResult = '-'
        
        return testResult

    def getTestedDevicesFromReport(self):
        existingDataInReport = pd.read_excel(self.reportSummary, sheet_name=self.mainSheetName, header=0)                         
        testedDevicesInReport = existingDataInReport['SERIAL_NUMBER']         
        self.testedDevices = [device for device in testedDevicesInReport if not pd.isnull(device)]

    def getParsedLogFilesFromReport(self):
        existingDataInReport = pd.read_excel(self.reportSummary, sheet_name=self.mainSheetName, header=0)                         
        parsedLogsInReport = existingDataInReport['Log_File_Name']         
        self.parsedLogs = [log for log in parsedLogsInReport if not pd.isnull(log)]
    
        # This is initialised in __initialiseReportSummary but if there is an existing report it must be calculated
        self.logFileNameColumn = len(existingDataInReport.columns)

    def getFirstEmptyRowFromReport(self):
        """ Retrieves the first empty row in the existing report, saves it as self.excelFirstEmptyRow"""
        
        workbook        = load_workbook(self.reportSummary)        
        worksheetIndex  = workbook.sheetnames.index(self.mainSheetName)
        worksheet       = workbook.worksheets[worksheetIndex]

        if self.dataOrder != "chronological":
            # Insert row at the top of the spreadsheet to write the data if anti-chronological order
            self.rowToWriteData = 2

        else:
            self.rowToWriteData = 1
                
        for col_cells in worksheet.iter_cols(min_col=1, max_col=1):
            for cell in col_cells:
                if cell.value is None:
                    break
                else:
                    self.rowToWriteData += 1
        
        # Close the workbook
        workbook.close()

    def highlightNonPerformedTestCellsRed(self, testInds):
        """ Highlights all empty cells RED where a test should have been performed but wasn't"""
        
        workbook        = load_workbook(self.reportSummary)        
        worksheetIndex  = workbook.sheetnames.index(self.mainSheetName)
        worksheet       = workbook.worksheets[worksheetIndex]
        
        # Remember: We add 3 to the index because we've left space for "Master Pass/Fail" and "Restest?" + indexing from 1 in Excel
        for ind in testInds:
            worksheet.cell(row = self.rowToWriteData, column = ind+3).fill = self.redFill
        
        workbook.save(self.reportSummary)
        workbook.close()

    def setPassFailCriteria(self):
        # Check that pass fail criterion document exists
        if not os.path.isfile('passFailCriteria/ballQAPassFailCriteria.xlsx'):
            sys.exit('Either the Pass Fail Criteria document or the folder does not exist, ')

        self.passFailDataFrame = pd.read_excel(r'passFailCriteria/ballQAPassFailCriteria.xlsx', sheet_name='Criteria Sheet')

        # Extract the test names and section headings
        self.__extractTestDetails(self.passFailDataFrame)

    def writePassFailCriteriaToExcel(self):
        """If the report does not exist we write the criteria into the freshly initialised file """
        
        workbook        = load_workbook(self.reportSummary)   
        workbook.create_sheet('Test Criteria')
        
        worksheetIndex  = workbook.sheetnames.index('Test Criteria')
        worksheet       = workbook.worksheets[worksheetIndex]
    
        # Write in the headers for the three columns
        worksheet.cell(row = 1, column = 1).value = 'Test Names'
        worksheet.cell(row = 1, column = 1).font = Font(bold = True, underline = 'single')
        worksheet.cell(row = 1, column = 1).fill = self.greenFill
        
        worksheet.cell(row = 1, column = 2).value = 'Minimum Value'
        worksheet.cell(row = 1, column = 2).font = Font(bold = True, underline = 'single')
        worksheet.cell(row = 1, column = 2).fill = self.greenFill
        
        worksheet.cell(row = 1, column = 3).value = 'Maximum Value'
        worksheet.cell(row = 1, column = 3).font = Font(bold = True, underline = 'single')
        worksheet.cell(row = 1, column = 3).fill = self.greenFill
                
        for j in range(len(self.testNames)):             
            # Write the data into the cell
            worksheet.cell(row = j+2, column = 1).value = self.testNames[j]            
            
            worksheet.cell(row = j+2, column = 2).value = self.testCriteria[j][0]
            worksheet.cell(row = j+2, column = 2).alignment = Alignment(horizontal = 'center')
            
            worksheet.cell(row = j+2, column = 3).value = self.testCriteria[j][1]
            worksheet.cell(row = j+2, column = 3).alignment = Alignment(horizontal = 'center')

        workbook.save(self.reportSummary)
        workbook.close()

    def writeDataToExcel(self, data, passFail, columns, file):
        """Write test results to Excel and use color formatting to show PASS/FAIL"""
        
        workbook        = load_workbook(self.reportSummary)        
        worksheetIndex  = workbook.sheetnames.index(self.mainSheetName)
        worksheet       = workbook.worksheets[worksheetIndex]
        
        if self.dataOrder != "chronological":
            #if worksheet.cell(row=2,column=1).value != None:
            worksheet.insert_rows(idx=2, amount = 1)

        for j, result in enumerate(data[0]):
            # Write the data into the cell
            worksheet.cell(row = self.rowToWriteData, column = j+1).value = result
            
            # Format the cell to show if it's a PASS/FAIL
            if isinstance(passFail[0][j], str):
                if passFail[0][j].lower() == 'pass':
                    worksheet.cell(row = self.rowToWriteData, column = j+1).fill = self.greenFill
                elif passFail[0][j].lower() == 'fail':
                    worksheet.cell(row = self.rowToWriteData, column = j+1).fill = self.redFill

        # Finally in the last column I'll add Log_File_Name so that when the code is run, we do not add previously parsed logs
        worksheet.cell(row = self.rowToWriteData, column = self.logFileNameColumn).value = file

        # Freeze first row
        worksheet.freeze_panes = 'A2'

        workbook.save(self.reportSummary)
        workbook.close()

    def writeReports(self):
        """ Write all the reports into the Excel file"""
        lst = []
        lst_folder = []
        latestTestOutcome = None 
        # Now loop over the folders and write the reports
        lst_folder = self.folders
        lst_folder.sort()         
        for folder in lst_folder:
            
            # Get rid of non-folder files that creep in
            if len(folder.split('.')) != 1:
                continue
            
            # Loop over the files
            lst = os.listdir( self.testLogFolder + '/' + folder)
            lst.sort()
            #logSortedList = sorted(lst, key=int)
            for file in tqdm(lst):
                
                if file.endswith(".log"):
                    
                    logFile = self.testLogFolder + '/' + folder + '/' + file
                    if logFile not in self.parsedLogs:
                        latestTestOutcome = self.createReportFor(logFile)
                        if self.dataOrder == "chronological":
                            self.rowToWriteData +=1 # Go to next line in Excel document
                        
                        # We may have duplicate log files erroneously in other folders
                        self.parsedLogs.append(logFile)
                        self.writeToLogsToInvestigate('\n')                    
                else:
                    continue
        
        return latestTestOutcome

    def writeToLogsToInvestigate(self, string):

        with open(self.logsToInvestigate, "a") as resultFile:                        
            resultFile.write(string)

    def createSummarySheet(self):
        """ Create Summary sheet in the report and states the outcome of each device that has been tested"""
        
        if self.testLogFolder == "@Valve":
            self.summarySheet = 'VALVE SUMMARY'
        elif self.testLogFolder == "@Bladder":
            self.summarySheet = 'BLADDER SUMMARY'
        elif self.testLogFolder == "@Ball":
            self.summarySheet = 'BALL SUMMARY'    

        #Initialize sheet
        try :
            workbook  = load_workbook(self.reportSummary)
            workbook.remove(workbook[self.summarySheet])
            workbook.save(self.reportSummary)
            workbook.close()   
        except :
            # Sheet does not exist
            None

        # Initializing Variables
        serialNumberHeader = "SERIAL_NUMBER"
        testedProduct = pd.DataFrame(columns=["SERIAL_NUMBER", "OUTCOME", "NB OF TESTS", "NB OF PASSES", "NB OF FAILS", "LATEST_DATE", "LATEST_TIME"])
        totalDevices = 0
        firstTimePasses = 0
        totalUnitPasses = 0
        totalTestPasses = 0
        totalTestTime = datetime.timedelta()
        timeCtr = 0
        fpyTable = []

        # Read Test Report
        masterReportData = pd.read_excel(self.reportSummary, self.mainSheetName, header=0)
        
        # Analyze test report and determine which devices passed or failed
        for line in masterReportData.index :
            if pd.isnull( masterReportData.at[ line, serialNumberHeader ] ) or pd.isnull( masterReportData.at[ line, "DATE" ] ) :
                continue

            #If the device has already been seen in the report
            if masterReportData.at[ line, serialNumberHeader ] in testedProduct["SERIAL_NUMBER"].to_list() :
                
                ind = testedProduct[ testedProduct["SERIAL_NUMBER"] == masterReportData.at[ line, serialNumberHeader ] ].index.values

                #Update the stats
                testedProduct.at[ ind[0], "NB OF TESTS" ]+=1

                if masterReportData.at[line, "Master Pass/Fail"] == "pass":
                    testedProduct.at[ ind[0], "NB OF PASSES" ]+=1
                    totalTestPasses+=1
                else :
                    testedProduct.at[ ind[0], "NB OF FAILS" ]+=1
                #If the time of the test is more recent
                if  masterReportData.at[ line, "DATE" ] > testedProduct.at[ ind[0], "LATEST_DATE"] :
                    #Update the time and the outcome
                    testedProduct.at[ ind[0], "OUTCOME" ] = masterReportData.at[ line, "Master Pass/Fail" ]
                    testedProduct.at[ ind[0], "LATEST_DATE" ] = masterReportData.at[ line, "DATE" ]
                    testedProduct.at[ ind[0], "LATEST_TIME" ] = masterReportData.at[ line, "START_TIME" ]
                elif masterReportData.at[ line, "DATE" ] == testedProduct.at[ ind[0], "LATEST_DATE"] and masterReportData.at[ line, "START_TIME" ] > testedProduct.at[ ind[0], "LATEST_TIME"] :
                    #Update the time and the outcome
                    testedProduct.at[ ind[0], "LATEST_DATE" ] = masterReportData.at[ line, "DATE" ]
                    testedProduct.at[ ind[0], "LATEST_TIME" ] = masterReportData.at[ line, "START_TIME" ]
                    testedProduct.at[ ind[0], "OUTCOME" ] = masterReportData.at[ line, "Master Pass/Fail" ]
                else :
                    if self.dataOrder != "chronological" :
                        if masterReportData.at[line, "Master Pass/Fail"] == "fail" and testedProduct.at[ ind[0], "OUTCOME" ] == "pass" :
                            if masterReportData.at[line, serialNumberHeader] not in fpyTable :
                                fpyTable.append( masterReportData.at[line, serialNumberHeader] )
                                firstTimePasses-=1 # If the previous test for that S/N was a fail
            else :
                totalDevices+=1
                if masterReportData.at[line, "Master Pass/Fail"] == "pass":
                    passIndicator = 1
                    firstTimePasses+=1
                    totalTestPasses+=1
                else :
                    passIndicator = 0

                df = pd.DataFrame({
                    "SERIAL_NUMBER" : [ masterReportData.at[line, serialNumberHeader] ],
                    "OUTCOME" : [ masterReportData.at[line, "Master Pass/Fail"] ],
                    "NB OF TESTS" : [1],
                    "NB OF PASSES" : [passIndicator],
                    "NB OF FAILS" : [ 1 - passIndicator],
                    "LATEST_DATE" : [ masterReportData.at[line, "DATE"] ],
                    "LATEST_TIME" : [ masterReportData.at[line, "START_TIME"] ]
                })

                if self.dataOrder != "chronological":
                    testedProduct = pd.concat( [testedProduct, df], ignore_index=True )
                else:
                    testedProduct = pd.concat( [df, testedProduct], ignore_index=True )

            if not pd.isnull( masterReportData.at[ line, "TEST_DURATION" ] ) :
                    testDuration = masterReportData.at[ line, "TEST_DURATION" ].split(":")
                    testDurationDelta = datetime.timedelta(hours=float(testDuration[0]), minutes=float(testDuration[1]), seconds=float(testDuration[2]) )
                    totalTestTime+=testDurationDelta
                    timeCtr+=1

        #Write dataframe to summary sheet
        with pd.ExcelWriter(self.reportSummary, engine="openpyxl", mode='a') as writer :
            testedProduct.to_excel(writer, sheet_name=self.summarySheet, index=False) 

        # Formatting sheet 
        workbook        = load_workbook(self.reportSummary)
        worksheetIndex  = workbook.sheetnames.index(self.summarySheet)
        worksheet       = workbook.worksheets[worksheetIndex]
        
        # Defining Dimensions
        worksheet.column_dimensions['A'].width = 18.5
        worksheet.column_dimensions['B'].width = 14.5
        worksheet.column_dimensions['C'].width = 14.5
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 14.5
        worksheet.column_dimensions['F'].width = 14.5
        worksheet.column_dimensions['G'].width = 14.5

        # Formatting header
        worksheet.cell(row = 1, column = 1).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 2).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 3).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 4).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 5).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 6).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 7).font = Font(name='Calibri', size=11, bold = True)
        
        for i in testedProduct.index:         
            # Format the cell to show if it's a PASS/FAIL
            if testedProduct.at[i, "OUTCOME"] == "pass" :
                worksheet.cell(row = i+2, column = 2).fill = self.greenFill
            elif testedProduct.at[i, "OUTCOME"] == "fail" :
                worksheet.cell(row = i+2, column = 2).fill = self.redFill

        for col in worksheet.iter_cols(min_col=1, max_col=7):
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        #Create yield table in the summary sheet
        worksheet.merge_cells('J4:N4')
        worksheet.merge_cells('J6:J7')
        worksheet.merge_cells('N6:N7')
        worksheet.merge_cells('O6:O7')

        worksheet['J4'] = "YIELD"
        worksheet['J5'] = "NB OF DEVICES"
        worksheet['K5'] = "PASSES"
        worksheet['L5'] = "FPY"
        worksheet['M5'] = "FAILS"
        worksheet['N5'] = "AVG TEST TIME (ALL)"
        worksheet['O5'] = "AVG TIME FOR A PASS"

        for result in testedProduct.index :
            if testedProduct.at[ result, "OUTCOME"] == "pass" :
                totalUnitPasses+=1

        if totalDevices > 0 :

            fpy = ( firstTimePasses / totalDevices ) * 100
            runYield = ( totalUnitPasses / totalDevices ) * 100
            
            worksheet['J6'] = totalDevices
            passes = (totalUnitPasses / totalDevices)*100
            worksheet['K6'] = f"{totalUnitPasses} / {totalDevices}"  
            worksheet['K7'] = f"{passes:.2f} %"
            worksheet['L6'] = f"{firstTimePasses} / {totalDevices}"
            fpy = (firstTimePasses / totalDevices)*100 
            worksheet['L7'] = f"{fpy:.2f} %"
                
            worksheet['M6'] = f"{totalDevices-totalUnitPasses} / {totalDevices}"
            fails = ((totalDevices-totalUnitPasses) / totalDevices)*100
            worksheet['M7'] = f"{fails:.2f} %"

            if timeCtr > 0 :
                avgTestTime = totalTestTime / timeCtr
                worksheet['N6'] = f"{avgTestTime.seconds//60}min{avgTestTime.seconds%60}secs"
            if totalTestPasses > 0 :
                avgTestTime = totalTestTime / totalTestPasses
                worksheet['O6'] = f"{avgTestTime.seconds//60}min{avgTestTime.seconds%60}secs"
        yieldTableHeaderFont = Font(name='Calibri', size=11,bold=True)
        worksheet['J4'].font = yieldTableHeaderFont
        worksheet['J4'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['J4'].border = Border(left = Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin') )
        worksheet['M4'].border = Border(right=Side(border_style='thin'))

        yieldTableTitlesFont = Font(name='Calibri', size=11,italic=True)
        worksheet['J5'].font = yieldTableTitlesFont
        worksheet['K5'].font = yieldTableTitlesFont
        worksheet['L5'].font = yieldTableTitlesFont
        worksheet['M5'].font = yieldTableTitlesFont
        worksheet['N5'].font = yieldTableTitlesFont
        worksheet['O5'].font = yieldTableTitlesFont

        for col in worksheet.iter_cols(min_col=10, max_col=15, min_row=5, max_row=8):
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in worksheet.iter_cols(min_col=10, max_col=15, min_row=4, max_row=7):
            for cell in col:
                cell.border = Border(left = Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin') )

        worksheet.column_dimensions['J'].width = 14.5
        worksheet.column_dimensions['M'].width = 13
        worksheet.column_dimensions['N'].width = 20
        worksheet.column_dimensions['O'].width = 24
        
        #Arrange the summary sheet orders
        if self.testLogFolder == "@Valve":
            workbook.move_sheet(self.summarySheet, offset=-5)
            workbook.move_sheet(self.mainSheetName, offset=-4)
        elif self.testLogFolder == "@Bladder":
            workbook.move_sheet(self.summarySheet, offset=-3)
            workbook.move_sheet(self.mainSheetName, offset=-2)
        elif self.testLogFolder == "@Ball":
            workbook.move_sheet(self.summarySheet, offset=-1)
            workbook.move_sheet(self.mainSheetName, offset=0)

        workbook.save(self.reportSummary)
        workbook.close()

    def createMasterSummarySheet(self):
        """ Create Master Summary Sheet in the case of multiple test stages"""    

        #Initialize sheet
        # Initializing Variables
        masterSummarySheetName = "MASTER SUMMARY"
        
        totalDevices = 0
        totalFullPasses = 0
        totalPartialPasses = 0    
        totalFails = 0
        totalTested = 0

        workbook  = load_workbook(self.reportSummary)
        if "MASTER SUMMARY" not in workbook.sheetnames:
            #masterSummarySheet = workbook.create_sheet(masterSummarySheetName)
            masterSummaryData = pd.DataFrame(columns=["SERIAL_NUMBER", "VALVE TEST", "TEST DATE 1", "BLADDER TEST", "TEST DATE 2", "BALL TEST", "TEST DATE 3", "FINAL OUTCOME"])
        else :
            masterSummaryData = pd.read_excel(self.reportSummary, masterSummarySheetName, header=0)
            masterSummaryData.drop(masterSummaryData.columns[[8,9,10,11,12,13]], axis=1, inplace=True)
            toScrape = []
            for i in masterSummaryData.index:
                if pd.isnull(masterSummaryData.at[i, "SERIAL_NUMBER"]) :
                    toScrape.append(i)
            if toScrape :
                masterSummaryData.drop(toScrape, axis=0, inplace=True)
            
            #workbook.remove(workbook[masterSummarySheetName])    
            with pd.ExcelWriter(self.reportSummary, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer :
                masterSummaryData.to_excel(writer, masterSummarySheetName,index=False) 


        workbook.save(self.reportSummary)
        workbook.close()   

        # Read Summary Sheet for the current test
        summaryReportData = pd.read_excel(self.reportSummary, self.summarySheet, header=0)
        
        # Analyze master report and determine which devices passed or failed
        for line in summaryReportData.index :
            if pd.isnull( summaryReportData.at[ line, "SERIAL_NUMBER" ] ) :
                continue
            
            #If the device has already been seen in the report
            if summaryReportData.at[ line, "SERIAL_NUMBER" ] in masterSummaryData["SERIAL_NUMBER"].to_list() :
                
                ind = masterSummaryData[ masterSummaryData["SERIAL_NUMBER"] == summaryReportData.at[ line, "SERIAL_NUMBER" ] ].index.values
                                    
                #Update the time and the outcome
                if self.mainSheetName == "Valve Report" :
                    masterSummaryData.at[ ind[0], "VALVE TEST" ] = summaryReportData.at[ line, "OUTCOME" ].upper()
                    masterSummaryData.at[ ind[0], "TEST DATE 1" ] = summaryReportData.at[ line, "LATEST_DATE" ]
                elif self.mainSheetName == "Bladder Report" :
                    masterSummaryData.at[ ind[0], "BLADDER TEST" ] = summaryReportData.at[ line, "OUTCOME" ].upper()
                    masterSummaryData.at[ ind[0], "TEST DATE 2" ] = summaryReportData.at[ line, "LATEST_DATE" ]
                elif self.mainSheetName == "Ball Report" :
                    masterSummaryData.at[ ind[0], "BALL TEST" ] = summaryReportData.at[ line, "OUTCOME" ].upper()
                    masterSummaryData.at[ ind[0], "TEST DATE 3" ] = summaryReportData.at[ line, "LATEST_DATE" ]
                
                if masterSummaryData.at[ ind[0], "VALVE TEST"] == "PASS" and masterSummaryData.at[ ind[0], "BLADDER TEST"] == "PASS" and masterSummaryData.at[ ind[0], "BALL TEST"] == "PASS" :
                    masterSummaryData.at[ ind[0], "FINAL OUTCOME" ] = "PASS"
                elif "PASS" in masterSummaryData.iloc[ind[0],1:4].tolist() and "FAIL" not in masterSummaryData.iloc[ind[0],1:4].tolist():
                    masterSummaryData.at[ ind[0], "FINAL OUTCOME" ] = "PARTIAL PASS"
                elif "FAIL" in masterSummaryData.iloc[ind[0],1:4].tolist() :
                    masterSummaryData.at[ ind[0], "FINAL OUTCOME" ] = "FAIL"
            
            else :

                if self.mainSheetName == "Valve Report":

                    df = pd.DataFrame({
                        "SERIAL_NUMBER" : [ summaryReportData.at[line, "SERIAL_NUMBER"] ],
                        "VALVE TEST" : [ summaryReportData.at[line, "OUTCOME"].upper() ],
                        "TEST DATE 1" : [ summaryReportData.at[line, "LATEST_DATE"] ],
                        "BLADDER TEST" : ["TBD"],
                        "TEST DATE 2" : ["-"],
                        "BALL TEST" : ["TBD"],
                        "TEST DATE 3" : ["-"],
                        "FINAL OUTCOME" : [ "PARTIAL PASS" if summaryReportData.at[line, "OUTCOME"].upper()=="PASS" else "FAIL" ],
                    })

                elif self.mainSheetName == "Bladder Report":

                    df = pd.DataFrame({
                        "SERIAL_NUMBER" : [ summaryReportData.at[line, "SERIAL_NUMBER"] ],
                        "VALVE TEST" : ["TBD"],
                        "TEST DATE 1" : ["-"],
                        "BLADDER TEST" : [ summaryReportData.at[line, "OUTCOME"].upper() ],
                        "TEST DATE 2" : [ summaryReportData.at[line, "LATEST_DATE"] ],
                        "BALL TEST" : ["TBD"],
                        "TEST DATE 3" : ["-"],
                        "FINAL OUTCOME" : [ "PARTIAL PASS" if summaryReportData.at[line, "OUTCOME"].upper()=="PASS" else "FAIL" ],
                    })

                elif self.mainSheetName == "Ball Report":

                    df = pd.DataFrame({
                        "SERIAL_NUMBER" : [ summaryReportData.at[line, "SERIAL_NUMBER"] ],
                        "VALVE TEST" : ["TBD"],
                        "TEST DATE 1" : ["-"],
                        "BLADDER TEST" : ["TBD"],
                        "TEST DATE 1" : ["-"],
                        "BALL TEST" : [ summaryReportData.at[line, "OUTCOME"].upper() ],
                        "TEST DATE 3" : [ summaryReportData.at[line, "LATEST_DATE"] ],
                        "FINAL OUTCOME" : ["PARTIAL PASS" if summaryReportData.at[line, "OUTCOME"].upper()=="PASS" else "FAIL"],
                    })
             
                masterSummaryData = pd.concat( [masterSummaryData,df], ignore_index=True )

        #Write dataframe to summary sheet
        with pd.ExcelWriter(self.reportSummary, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer :
            masterSummaryData.to_excel(writer, masterSummarySheetName,index=False) 

        # Formatting sheet 
        workbook        = load_workbook(self.reportSummary)
        worksheetIndex  = workbook.sheetnames.index(masterSummarySheetName)
        worksheet       = workbook.worksheets[worksheetIndex]
        
        # Defining Dimensions
        worksheet.column_dimensions['A'].width = 18.5
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 16
        worksheet.column_dimensions['D'].width = 17
        worksheet.column_dimensions['E'].width = 16
        worksheet.column_dimensions['F'].width = 15
        worksheet.column_dimensions['G'].width = 16
        worksheet.column_dimensions['H'].width = 18
        
        # Formatting header
        worksheet.cell(row = 1, column = 1).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 2).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 3).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 4).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 5).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 6).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 7).font = Font(name='Calibri', size=11, bold = True)
        worksheet.cell(row = 1, column = 8).font = Font(name='Calibri', size=11, bold = True)
        
        for i in masterSummaryData.index:         
            # Format the cell to show if it's a PASS/FAIL
            for j in range(1,8):

                if not pd.isnull( masterSummaryData.iloc[i, j] ):

                    if masterSummaryData.iloc[i, j] == "PASS" :
                        worksheet.cell(row = i+2, column = j+1).fill = self.greenFill
                    elif masterSummaryData.iloc[i, j] == "PARTIAL PASS" :
                        worksheet.cell(row = i+2, column = j+1).fill = self.yellowFill
                    elif masterSummaryData.iloc[i, j] == "FAIL" :
                        worksheet.cell(row = i+2, column = j+1).fill = self.redFill

        for col in worksheet.iter_cols(min_col=1, max_col=8):
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        #Create yield table in the summary sheet
        worksheet.merge_cells('J4:N4')
        worksheet.merge_cells('J6:J7')

        worksheet['J4'] = "YIELD"
        worksheet['J5'] = "NB OF DEVICES"
        worksheet['K5'] = "FULLY TESTED"
        worksheet['L5'] = "FULL PASSES"
        worksheet['M5'] = "PARTIAL PASSES"
        worksheet['N5'] = "FAILS"
            
        for result in masterSummaryData.index :
            
            totalDevices+=1
            if masterSummaryData.at[ result, "FINAL OUTCOME"] == "PASS" :
                totalFullPasses+=1
            
            if masterSummaryData.at[ result, "FINAL OUTCOME"] == "PARTIAL PASS" :
                totalPartialPasses+=1
            
            if masterSummaryData.at[ result, "FINAL OUTCOME"] == "FAIL" :
                totalFails+=1
            
            if "TBD" not in masterSummaryData.iloc[ result, 1:4].tolist() :
                totalTested+=1
            
        if totalDevices > 0 :
            
            worksheet['J6'] = totalDevices

            fullyTested = (totalTested/ totalDevices)*100
            worksheet['K6'] = f"{totalTested} / {totalDevices}"  
            worksheet['K7'] = f"{fullyTested:.2f} %"

            fullPasses = (totalFullPasses / totalDevices)*100
            worksheet['L6'] = f"{totalFullPasses} / {totalDevices}"  
            worksheet['L7'] = f"{fullPasses:.2f} %"

            partialPasses = (totalPartialPasses / totalDevices)*100
            worksheet['M6'] = f"{totalPartialPasses} / {totalDevices}"  
            worksheet['M7'] = f"{partialPasses:.2f} %"
                
            worksheet['N6'] = f"{totalFails} / {totalDevices}"
            fails = ((totalFails) / totalDevices)*100
            worksheet['N7'] = f"{fails:.2f} %"

        yieldTableHeaderFont = Font(name='Calibri', size=11,bold=True)
        worksheet['J4'].font = yieldTableHeaderFont
        worksheet['J4'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['J4'].border = Border(left = Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin') )
        worksheet['M4'].border = Border(right=Side(border_style='thin'))

        yieldTableTitlesFont = Font(name='Calibri', size=11,italic=True)
        worksheet['J5'].font = yieldTableTitlesFont
        worksheet['K5'].font = yieldTableTitlesFont
        worksheet['L5'].font = yieldTableTitlesFont

        for col in worksheet.iter_cols(min_col=10, max_col=14, min_row=5, max_row=7):
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in worksheet.iter_cols(min_col=10, max_col=14, min_row=4, max_row=7):
            for cell in col:
                cell.border = Border(left = Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin') )

        worksheet.column_dimensions['J'].width = 15
        worksheet.column_dimensions['K'].width = 15
        worksheet.column_dimensions['L'].width = 15
        worksheet.column_dimensions['M'].width = 16
        worksheet.column_dimensions['N'].width = 15
        
        #Arrange the summary sheet orders
        sheets = workbook._sheets.copy()
        sheets.insert(0, sheets.pop(sheets.index(workbook["MASTER SUMMARY"])))
        sheets.insert(1, sheets.pop(sheets.index(workbook[self.summarySheet])))
        sheets.insert(2, sheets.pop(sheets.index(workbook[self.mainSheetName])))
        workbook._sheets = sheets

        workbook.save(self.reportSummary)
        workbook.close()

    def backupProductionFolder(self, verbosity, architecture):
        """ Add logs and report from the current test folder to the shared production drive"""
        
        # Get the user name
        username = os.getenv("USER")
        if username == "root" :
            username = os.getenv("SUDO_USER")

        homePath = "/home/" + username
        driveFolder = homePath + '/Desktop/sportableTestResult/'
        lst = os.listdir(driveFolder)
        #productFolder = self.testLogFolder.strip('@')
        productFolder = "SMART BALL"
        batchFolder = ''

        #Get batch folder name
        batchFolder = os.getcwd().split('/')

        if batchFolder[-1] == 'release' :
            batchFolder = batchFolder[-2]
        else :
            batchFolder = batchFolder[-1]

        try:
            if architecture == "singleStage" :
                for i in range(len(lst)):
                    if productFolder.upper() in lst[i]:
                        productFolder = lst[i]

                #Check if the batch folder exists in the drive
                if batchFolder not in os.listdir(driveFolder + "/" + productFolder + "/"):
                    raise Exception(batchFolder + " folder not found on the drive.\nPlease check the folder name or the drive.")

                sharedBatchFolder = driveFolder + productFolder + "/" + batchFolder + "/Build test data/Electrical test data/"
            
            if architecture == "multiStage" :
                sharedBatchFolder = driveFolder + batchFolder + "/"
                subprocess.call(['sudo', 'cp', self.reportSummary, homePath + '/Desktop/'])
                subprocess.call(['sudo', 'chmod', 'ugo=r', homePath + '/Desktop/' + self.reportSummary])

        except Exception as error:
            print(" /!\ ERROR ACCESSING THE FOLDER IN THE SHARED DRIVE /!\ ")
            print(" ==> " + str(error))
            exit()

        try :

        #Copy recursively the folders and log non-existing in the shared Folder
            print("Uploading log to the shared drive")
            if verbosity == "quiet" :
                subprocess.call(['rclone', 'copy', '-q', self.testLogFolder + '/', sharedBatchFolder + self.testLogFolder])
                subprocess.call(['rclone', 'copy', '-q', self.reportSummary , sharedBatchFolder])
            elif verbosity == "debug" :
                subprocess.call(['rclone', 'copy', '-vP', self.testLogFolder, sharedBatchFolder + self.testLogFolder])
                subprocess.call(['rclone', 'copy', '-vP', self.reportSummary , sharedBatchFolder])

        except Exception as error:
            print(" /!\ ERROR UPLOADING THE DATA TO THE SHARED DRIVE /!\ ")
            print(" ==> " + str(error))
            exit()
    
#%% Run main

if __name__ == "__main__":

    productionConfig = configReader.ProductionConfig("assets/productionConfig.json")
    if "--chronological" in productionConfig.commandLineArguments :
        dataOrder = "chronological"
    else :
        dataOrder = ""

    if "--valve-test" in productionConfig.commandLineArguments:
        stage = "Valve"
    elif "--bladder-test" in productionConfig.commandLineArguments:
        stage = "Bladder"
    elif "--ball-test" in productionConfig.commandLineArguments:
        stage = "Ball"


    # Initialise the class
    brg = ballReportGenerator(stage, dataOrder)
    
    # Write the reports
    brg.writeReports()

    brg.createSummarySheet()

    if "--master-summary" in productionConfig.commandLineArguments:
        brg.createMasterSummarySheet()
        brg.backupProductionFolder("debug", "multiStage")
    else :
        brg.backupProductionFolder("debug", "singleStage")

    # Make it read-only and open the spreadsheet
    os.system("libreoffice --calc --view *Electrical_Test_Report.xlsx &")
